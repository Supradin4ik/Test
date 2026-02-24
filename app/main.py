from pathlib import Path

from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi import Request
from openpyxl import load_workbook

from app.db import get_conn, init_db
from app.pdf_index import ensure_pdf_schema, find_pdf, reindex_source

app = FastAPI(title="Workshop Flow")
app.mount("/static", StaticFiles(directory="app"), name="static")
templates = Jinja2Templates(directory="app/templates")


@app.on_event("startup")
def startup() -> None:
    init_db()
    conn = get_conn()
    ensure_pdf_schema(conn)
    conn.close()


def read_settings(conn):
    return conn.execute("SELECT * FROM project_settings WHERE id = 1").fetchone()


@app.get("/", response_class=HTMLResponse)
def project_page(request: Request):
    conn = get_conn()
    settings = read_settings(conn)
    missing = conn.execute("SELECT COUNT(*) AS c FROM parts WHERE pdf_missing = 1").fetchone()["c"]
    conn.close()
    return templates.TemplateResponse("project.html", {"request": request, "settings": settings, "missing": missing})


@app.post("/reindex")
def reindex(project_root_path: str = Form(...), standard_root_path: str = Form(...)):
    conn = get_conn()
    reindex_source(conn, "project", project_root_path)
    reindex_source(conn, "standard", standard_root_path)
    conn.close()
    return RedirectResponse(url="/", status_code=303)


@app.post("/generate")
def generate_orders(
    excel_file: UploadFile = File(...),
    project_root_path: str = Form(...),
    standard_root_path: str = Form(...),
    daily_kits: int = Form(...),
    total_kits: int = Form(...),
):
    upload_dir = Path("uploads")
    upload_dir.mkdir(exist_ok=True)
    excel_path = upload_dir / excel_file.filename
    excel_path.write_bytes(excel_file.file.read())

    conn = get_conn()
    reindex_source(conn, "project", project_root_path)
    reindex_source(conn, "standard", standard_root_path)

    conn.execute("DELETE FROM parts")
    conn.execute("DELETE FROM project_settings WHERE id = 1")
    conn.execute(
        "INSERT INTO project_settings(id, project_root_path, standard_root_path, daily_kits, total_kits, excel_path) VALUES(1, ?, ?, ?, ?, ?)",
        (project_root_path, standard_root_path, daily_kits, total_kits, str(excel_path)),
    )

    wb = load_workbook(excel_path, data_only=True)
    sheet = wb["ORDER"]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or row[1] is None:
            continue
        position = str(row[0] or "").strip()
        dxf_name = str(row[1] or "").strip()
        part_name = str(row[2] or "").strip()
        qty_per_kit = float(row[3] or 0)
        thickness = str(row[4] or "").strip()
        material = str(row[5] or "").strip()
        qty_day = qty_per_kit * daily_kits

        pdf_path = find_pdf(conn, dxf_name, "project") or find_pdf(conn, dxf_name, "standard")
        missing = 1 if not pdf_path else 0

        conn.execute(
            """
            INSERT INTO parts(position, dxf_name, part_name, qty_per_kit, thickness, material, qty_day, pdf_path, pdf_missing)
            VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (position, dxf_name, part_name, qty_per_kit, thickness, material, qty_day, pdf_path, missing),
        )
    conn.commit()
    conn.close()
    return RedirectResponse(url="/laser", status_code=303)


@app.get("/laser", response_class=HTMLResponse)
def laser_page(request: Request):
    conn = get_conn()
    settings = read_settings(conn)
    groups = conn.execute(
        """
        SELECT material, thickness, COUNT(*) as items,
               SUM(qty_day) as qty,
               SUM(CASE WHEN laser_status = 'cut_done' THEN 1 ELSE 0 END) as done
        FROM parts
        GROUP BY material, thickness
        ORDER BY material, thickness
        """
    ).fetchall()

    details = conn.execute("SELECT * FROM parts ORDER BY material, thickness, dxf_name").fetchall()
    conn.close()
    return templates.TemplateResponse("laser.html", {"request": request, "groups": groups, "parts": details, "settings": settings})


@app.post("/laser/{part_id}/status")
def update_laser_status(part_id: int, status: str = Form(...), zone: str = Form("")):
    conn = get_conn()
    if status not in {"in_progress", "cut_done"}:
        raise HTTPException(400, "Bad status")
    conn.execute("UPDATE parts SET laser_status = ?, zone = CASE WHEN ? != '' THEN ? ELSE zone END WHERE id = ?", (status, zone, zone, part_id))
    if status == "cut_done":
        conn.execute("UPDATE parts SET bend_status = 'ready' WHERE id = ?", (part_id,))
    conn.commit()
    conn.close()
    return RedirectResponse(url="/laser", status_code=303)


@app.post("/stage/{stage}/{part_id}")
def update_stage(stage: str, part_id: int, status: str = Form(...)):
    transitions = {
        "bend": ("bend_status", "weld_status"),
        "weld": ("weld_status", "qc_status"),
        "qc": ("qc_status", None),
    }
    if stage not in transitions:
        raise HTTPException(404, "Unknown stage")
    column, next_column = transitions[stage]
    conn = get_conn()
    conn.execute(f"UPDATE parts SET {column} = ? WHERE id = ?", (status, part_id))
    if status == "done" and next_column:
        conn.execute(f"UPDATE parts SET {next_column} = 'ready' WHERE id = ?", (part_id,))
    conn.commit()
    conn.close()
    return RedirectResponse(url=f"/{stage}", status_code=303)


@app.get("/{stage}", response_class=HTMLResponse)
def stage_page(stage: str, request: Request):
    if stage not in {"bend", "weld", "qc"}:
        raise HTTPException(404, "not found")
    column = f"{stage}_status"
    conn = get_conn()
    rows = conn.execute(
        f"SELECT * FROM parts WHERE {column} IN ('ready','in_progress') ORDER BY material, thickness, dxf_name"
    ).fetchall()
    conn.close()
    return templates.TemplateResponse("stage.html", {"request": request, "stage": stage, "rows": rows, "column": column})


@app.get("/pdf/{part_id}")
def open_pdf(part_id: int):
    conn = get_conn()
    row = conn.execute("SELECT pdf_path FROM parts WHERE id = ?", (part_id,)).fetchone()
    conn.close()
    if not row or not row["pdf_path"]:
        raise HTTPException(404, "PDF missing")
    return FileResponse(row["pdf_path"], media_type="application/pdf")


@app.get("/issues", response_class=HTMLResponse)
def issues(request: Request):
    conn = get_conn()
    rows = conn.execute("SELECT position, dxf_name, part_name FROM parts WHERE pdf_missing = 1 ORDER BY dxf_name").fetchall()
    conn.close()
    return templates.TemplateResponse("issues.html", {"request": request, "rows": rows})
